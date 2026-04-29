import cv2
import os
import sys
import winsound
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ---------- PATH FIX ----------
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath("")
    return os.path.join(base_path, relative_path)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------- MODEL ----------
model_path = resource_path("model.yml")

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read(model_path)

# ---------- LABELS (ADD MORE USERS HERE) ----------
labels = {
    0: "solan",
    1: "person2",
    2: "person3"
}

# ---------- CASCADE ----------
cascade_path = os.path.join(BASE_DIR, "haarcascade_frontalface_default.xml")
face_cascade = cv2.CascadeClassifier(cascade_path)

# ---------- CAMERA ----------
cap = cv2.VideoCapture(0)

# ---------- EXCEL FILE ----------
excel_path = os.path.join(BASE_DIR, "attendance.xlsx")

if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Time", "Status"])
    wb.save(excel_path)

# ---------- TRACKING ----------
last_marked_time = {}
status_text = ""
status_timer = 0

# ---------- MARK ATTENDANCE ----------
def mark_attendance(name):
    global status_text, status_timer

    now = datetime.now()
    time_string = now.strftime("%H:%M:%S")

    if name in last_marked_time:
        diff = (now - last_marked_time[name]).seconds
        if diff < 30:
            return False

    last_marked_time[name] = now

    wb = load_workbook(excel_path)
    ws = wb.active
    ws.append([name, time_string, "Present"])
    wb.save(excel_path)

    # 🔊 Sound
    winsound.Beep(1000, 300)

    # ✅ Status message
    status_text = f"{name} Marked ✅"
    status_timer = 30

    return True

# ---------- MAIN LOOP ----------
while True:
    success, img = cap.read()
    if not success:
        print("Camera not working")
        break

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    faces = face_cascade.detectMultiScale(gray, 1.3, 5)

    # ---------- UI ----------
    cv2.putText(img, "Face Attendance System", (10, 30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)

    cv2.putText(img, "Press Q to Exit", (10, 60),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

    cv2.putText(img, f"Faces: {len(faces)}", (10, 90),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

    # ---------- STATUS DISPLAY ----------
    if status_timer > 0:
        cv2.putText(img, status_text, (10, 120),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)
        status_timer -= 1

    # ---------- FACE LOOP ----------
    for (x, y, w, h) in faces:
        face = gray[y:y+h, x:x+w]

        try:
            id, confidence = recognizer.predict(face)
        except:
            continue

        if confidence < 70:
            name = labels.get(id, "Unknown")

            # GREEN BOX
            cv2.rectangle(img, (x, y), (x+w, y+h), (0, 255, 0), 2)

            # NAME + CONFIDENCE
            cv2.putText(img, f"{name} ({int(confidence)})", (x, y-10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

            mark_attendance(name)

        else:
            # RED BOX
            cv2.rectangle(img, (x, y), (x+w, y+h), (0, 0, 255), 2)

            cv2.putText(img, "Unknown", (x, y-10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)

    cv2.imshow("Attendance System", img)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# ---------- CLEANUP ----------
cap.release()
cv2.destroyAllWindows()

# ---------- AUTO OPEN EXCEL ----------
os.startfile(excel_path)