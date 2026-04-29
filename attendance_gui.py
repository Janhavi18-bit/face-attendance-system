import cv2
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ---------- MODEL ----------
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("model.yml")

# ---------- LABELS ----------
labels = {}
with open("labels.txt", "r") as f:
    for line in f:
        id, name = line.strip().split(",")
        labels[int(id)] = name

# ---------- FACE DETECTOR ----------
cascade_path = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
face_cascade = cv2.CascadeClassifier(cascade_path)

if face_cascade.empty():
    print("❌ Haarcascade not loaded")
    exit()

# ---------- CAMERA (IMPORTANT FIX) ----------
cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)

# ---------- EXCEL ----------
today = datetime.now().strftime("%Y-%m-%d")
excel_file = "Attendance.xlsx"

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Time", "Date"])
    wb.save(excel_file)

wb = load_workbook(excel_file)
ws = wb.active

marked = set()

print("🚀 Attendance System Running... Press Q to exit")

while True:
    ret, img = cap.read()
    if not ret:
        print("❌ Camera not detected")
        break

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    faces = face_cascade.detectMultiScale(gray, 1.3, 5)

    cv2.putText(img, "FACE ATTENDANCE SYSTEM", (10, 30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,255,255), 2)

    for (x, y, w, h) in faces:
        face = gray[y:y+h, x:x+w]

        try:
            id, confidence = recognizer.predict(face)
        except:
            continue

        if confidence < 70:
            name = labels.get(id, "Unknown")

            cv2.rectangle(img, (x,y), (x+w,y+h), (0,255,0), 2)
            cv2.putText(img, name, (x, y-10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,255,0), 2)

            if name not in marked:
                time_now = datetime.now().strftime("%H:%M:%S")
                ws.append([name, time_now, today])
                wb.save(excel_file)
                marked.add(name)

                print(f"✅ Marked: {name}")

        else:
            cv2.rectangle(img, (x,y), (x+w,y+h), (0,0,255), 2)
            cv2.putText(img, "Unknown", (x, y-10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,0,255), 2)

    cv2.imshow("Attendance System", img)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
wb.save(excel_file)
print("✅ Attendance Saved!")